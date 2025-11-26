import os
import sys
import pandas as pd
import cv2
import numpy as np
import io
import datetime
from itertools import combinations
from flask import Flask, render_template, request, session, Response
from werkzeug.utils import secure_filename
from skimage.metrics import structural_similarity as ssim
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
import re

# ---------------- CONFIG ----------------
PROJECT_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
sys.path.append(PROJECT_ROOT)

app = Flask(__name__, static_folder=os.path.join(PROJECT_ROOT, 'static'))
app.secret_key = 'shashank'

# Define and create necessary folder paths
UPLOAD_FOLDER = os.path.join(PROJECT_ROOT, 'uploads')
BRAND_IMAGES_FOLDER = os.path.join(PROJECT_ROOT, 'src', 'images')
DOCUMENTS_FOLDER = os.path.join(PROJECT_ROOT, 'src', 'documents')
DATA_FOLDER = os.path.join(PROJECT_ROOT, 'data')
for folder in [UPLOAD_FOLDER, BRAND_IMAGES_FOLDER, DOCUMENTS_FOLDER, DATA_FOLDER]:
    os.makedirs(folder, exist_ok=True)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# Master Data File Paths
EXPORT_DATA_PATH = os.path.join(DATA_FOLDER, 'export_export_data_filled_smart.csv')


# ---------------- HELPERS ----------------

def parse_csv_flexible(file_path):
    """Reads a CSV, trying 'utf-8' then 'latin1' encoding for robustness."""
    try:
        df = pd.read_csv(file_path, encoding='utf-8')
    except UnicodeDecodeError:
        df = pd.read_csv(file_path, encoding='latin1')
    except Exception as e:
        print(f"Error parsing CSV file '{os.path.basename(file_path)}': {e}")
        return pd.DataFrame()
    df.columns = df.columns.str.strip()
    return df


def compute_histogram(image_path):
    try:
        img_cv = cv2.imread(image_path)
        if img_cv is None: return None
        img_cv = cv2.resize(img_cv, (256, 256))
        hist = cv2.calcHist([img_cv], [0, 1, 2], None, [8, 8, 8], [0, 256, 0, 256, 0, 256])
        return cv2.normalize(hist, hist).flatten().astype('float32')
    except Exception:
        return None


def compute_ssim_score(image1_path, image2_path):
    try:
        img1 = cv2.imread(image1_path);
        img2 = cv2.imread(image2_path)
        if img1 is None or img2 is None: return None
        img1 = cv2.resize(img1, (256, 256));
        img2 = cv2.resize(img2, (256, 256))
        gray1 = cv2.cvtColor(img1, cv2.COLOR_BGR2GRAY);
        gray2 = cv2.cvtColor(img2, cv2.COLOR_BGR2GRAY)
        score, _ = ssim(gray1, gray2, full=True)
        return round(float(score), 4)
    except Exception:
        return None


def get_top_similar_docs(uploaded_path, docs_folder_path):
    try:
        with open(uploaded_path, 'r', encoding='utf-8', errors='ignore') as f:
            uploaded_text = f.read()
        return [("sample_proprietary_doc.txt", 0.91)] if "proprietary" in uploaded_text.lower() else [
            ("generic_manual.txt", 0.55)]
    except Exception:
        return []


# ---------------- ROUTES ----------------

@app.route('/')
def index():
    try:
        brand_folders = sorted(
            [f for f in os.listdir(BRAND_IMAGES_FOLDER) if os.path.isdir(os.path.join(BRAND_IMAGES_FOLDER, f))])
    except FileNotFoundError:
        return render_template('index.html', brand_folders=[], doc_folders=[], error="❌ Image folders not found.")
    doc_folder_path = os.path.join(PROJECT_ROOT, 'src', 'documents')
    doc_folders = sorted(
        [f for f in os.listdir(doc_folder_path) if os.path.isdir(os.path.join(doc_folder_path, f))]) if os.path.exists(
        doc_folder_path) else []
    return render_template('index.html', brand_folders=brand_folders, doc_folders=doc_folders)


@app.route('/submit_all', methods=['POST'])
def submit_all():
    """Handles file uploads and runs all analysis checks."""
    all_results = {'bom': {'high': [], 'low': []}, 'image': {'high': [], 'moderate': [], 'low': []},
                   'doc': {'high': [], 'low': []}, 'internal_sim': []}

    # --- BOM CHECK ---
    bom_file = request.files.get('bom_file')
    if bom_file and bom_file.filename.endswith('.csv'):
        filename = secure_filename(bom_file.filename)
        bom_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        bom_file.save(bom_path)
        bom_df = parse_csv_flexible(bom_path)

        try:
            export_df = parse_csv_flexible(EXPORT_DATA_PATH)
        except FileNotFoundError:
            all_results['bom']['high'].append({
                "Product": "Configuration Error",
                "Risk Level": "High",
                "Finding": "Master export data file is missing."
            })
            bom_df = pd.DataFrame()

        if not bom_df.empty:
            for _, row in bom_df.iterrows():
                category = str(row.get('Category', '')).strip()
                hs_code = str(row.get('HS Code', '')).strip()

                row_data = {
                    "Category": category,
                    "HS Code": hs_code,
                    "Company": row.get("Company", "N/A"),
                    "Product": row.get("Product", "N/A")
                }

                exact_match = export_df[
                    (export_df['Category'].str.strip().str.lower() == category.lower()) &
                    (export_df['HS Code'].astype(str).str.strip() == hs_code)
                    ]

                if not exact_match.empty:
                    row_data.update({
                        "Risk Level": "High",
                        "Finding": "Category and HS Code both found in export list."
                    })
                    all_results['bom']['high'].append(row_data)
                    continue

                hs_match = export_df[export_df['HS Code'].astype(str).str.strip() == hs_code]
                cat_match = export_df[export_df['Category'].str.strip().str.lower() == category.lower()]

                if not hs_match.empty or not cat_match.empty:
                    row_data.update({
                        "Risk Level": "Moderate",
                        "Finding": "Either Category or HS Code matches, but not both."
                    })
                    all_results['bom'].setdefault('moderate', []).append(row_data)
                    continue

                desc_match = any(category.lower() in str(exp_desc).lower() for exp_desc in export_df['Product Description'].dropna())
                if desc_match:
                    row_data.update({
                        "Risk Level": "Low",
                        "Finding": "Partial description similarity found in export list."
                    })
                    all_results['bom']['low'].append(row_data)
                else:
                    row_data.update({
                        "Risk Level": "Low",
                        "Finding": "No strong match, only weak/description similarity."
                    })
                    all_results['bom']['low'].append(row_data)

    # --- IMAGE CHECK ---
    image_file = request.files.get('image_file')
    if image_file and image_file.filename != '':
        filename = secure_filename(image_file.filename)
        save_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        image_file.save(save_path)
        uploaded_image_name = os.path.splitext(filename)[0].replace('_', ' ').strip()
        uploaded_hist = compute_histogram(save_path)
        if uploaded_hist is not None:
            selected_brand = request.form.get('brand_folder', 'all')
            brand_folders = [selected_brand] if selected_brand != 'all' else sorted(
                [d for d in os.listdir(BRAND_IMAGES_FOLDER) if os.path.isdir(os.path.join(BRAND_IMAGES_FOLDER, d))])
            image_match_found = False
            for brand in brand_folders:
                brand_images_path = os.path.join(BRAND_IMAGES_FOLDER, brand)
                for brand_image_name in os.listdir(brand_images_path):
                    if not brand_image_name.lower().endswith(('.png', '.jpg', '.jpeg')): continue
                    brand_image_path = os.path.join(brand_images_path, brand_image_name)
                    if os.path.isfile(brand_image_path):
                        brand_hist = compute_histogram(brand_image_path)
                        if brand_hist is None: continue
                        similarity_score = cv2.compareHist(uploaded_hist, brand_hist, cv2.HISTCMP_CORREL)
                        ssim_score = compute_ssim_score(save_path, brand_image_path)
                        risk_level = "Low"
                        if similarity_score > 0.85 or (ssim_score is not None and ssim_score > 0.80):
                            risk_level = "High"
                        elif similarity_score > 0.65 or (ssim_score is not None and ssim_score > 0.60):
                            risk_level = "Moderate"
                        if risk_level != "Low":
                            image_match_found = True
                            ssim_text = f"{ssim_score:.2f}" if ssim_score is not None else "N/A"
                            finding = f"{risk_level} visual similarity (Correlation: {similarity_score:.2f}, SSIM: {ssim_text})"
                            category = brand
                            company = os.path.splitext(brand_image_name)[0].replace('_', ' ').strip()
                            all_results['image'][risk_level.lower()].append(
                                {"Uploaded Image": uploaded_image_name, "Brand Image": brand_image_name,
                                 "Risk Level": risk_level, "Finding": finding, "Category": category,
                                 "Company": company})
            if not image_match_found:
                all_results['image']['low'].append({"Uploaded Image": uploaded_image_name, "Risk Level": "Low",
                                                    "Finding": "No significant visual similarity."})

    # --- CONSOLIDATION STEP ---
    # --- CONSOLIDATION STEP ---
    high_risk_bom_items = all_results['bom']['high']
    high_risk_image_items = all_results['image']['high']
    high_risk_doc_items = all_results['doc']['high']

    bom_items_to_remove, image_items_to_remove, doc_items_to_remove = [], [], []

    # 🔹 BOM ↔ Image cross validation
    for bom_item in high_risk_bom_items:
        bom_words = set(re.sub(r'[\W_]+', ' ', bom_item.get('Category', '')).lower().split())
        for image_item in high_risk_image_items:
            image_words = set(re.sub(r'[\W_]+', ' ', image_item.get('Category', '')).lower().split())
            if bom_words and image_words and bom_words.intersection(image_words):
                all_results['internal_sim'].append({
                    "Type": "BOM-Image Cross-Validation",
                    "Risk Level": "High",
                    "Product": bom_item.get('Product'),
                    "Category": bom_item.get('Category'),
                    "Company": bom_item.get('Company'),
                    "Finding": f"BOM high-risk item confirmed by Image (match: {image_item.get('Brand Image')})."
                })
                bom_items_to_remove.append(bom_item)
                image_items_to_remove.append(image_item)
                break

    # 🔹 BOM ↔ Document cross validation
    for bom_item in high_risk_bom_items:
        bom_words = set(re.sub(r'[\W_]+', ' ', bom_item.get('Category', '')).lower().split())
        for doc_item in high_risk_doc_items:
            doc_words = set(re.sub(r'[\W_]+', ' ', doc_item.get('Category', '')).lower().split())
            if bom_words and doc_words and bom_words.intersection(doc_words):
                all_results['internal_sim'].append({
                    "Type": "BOM-Document Cross-Validation",
                    "Risk Level": "High",
                    "Product": bom_item.get('Product'),
                    "Category": bom_item.get('Category'),
                    "Company": bom_item.get('Company'),
                    "Finding": f"BOM high-risk item confirmed by Document (doc: {doc_item.get('Finding')})."
                })
                bom_items_to_remove.append(bom_item)
                doc_items_to_remove.append(doc_item)
                break

    # 🔹 Image ↔ Document cross validation
    for image_item in high_risk_image_items:
        image_words = set(re.sub(r'[\W_]+', ' ', image_item.get('Category', '')).lower().split())
        for doc_item in high_risk_doc_items:
            doc_words = set(re.sub(r'[\W_]+', ' ', doc_item.get('Category', '')).lower().split())
            if image_words and doc_words and image_words.intersection(doc_words):
                all_results['internal_sim'].append({
                    "Type": "Image-Document Cross-Validation",
                    "Risk Level": "High",
                    "Product": image_item.get('Uploaded Image'),
                    "Category": image_item.get('Category'),
                    "Company": image_item.get('Company'),
                    "Finding": f"Image high-risk item confirmed by Document (doc: {doc_item.get('Finding')})."
                })
                image_items_to_remove.append(image_item)
                doc_items_to_remove.append(doc_item)
                break

    # Remove confirmed items from standalone lists
    all_results['bom']['high'] = [i for i in high_risk_bom_items if i not in bom_items_to_remove]
    all_results['image']['high'] = [i for i in high_risk_image_items if i not in image_items_to_remove]
    all_results['doc']['high'] = [i for i in high_risk_doc_items if i not in doc_items_to_remove]

    session['analysis_results'] = all_results
    return render_template('results.html', results=all_results)


@app.route('/generate_report', methods=['GET'])
def generate_report():
    """Generates the consolidated Excel report."""
    all_results = session.get('analysis_results', {})
    if not all_results:
        return Response("No analysis results found.", mimetype='text/plain', status=404)

    wb = Workbook()
    ws = wb.active
    ws.title = "IP Risk Report"
    headers = ["Analysis Type", "Risk Level", "Category", "Product", "Company", "Finding"]
    ws.append(headers)
    for cell in ws[1]: cell.font = Font(bold=True)

    red_fill = PatternFill(start_color="FFC7CE", fill_type="solid")
    orange_fill = PatternFill(start_color="FFEB9C", fill_type="solid")
    green_fill = PatternFill(start_color="C6EFCE", fill_type="solid")
    blue_fill = PatternFill(start_color="DDEBF7", fill_type="solid")

    def get_risk_fill(risk_level):
        rl = str(risk_level).lower()
        if "high" in rl: return red_fill
        if "moderate" in rl: return orange_fill
        if "low" in rl: return green_fill
        return None

    # 1. Print Combined (Cross-Validation) Results FIRST
    for item in all_results.get('internal_sim', []):
        row = [item.get('Type'), item.get('Risk Level'), item.get('Category'), item.get('Product'), item.get('Company'),
               item.get('Finding')]
        ws.append(row)
        for cell in ws[ws.max_row]: cell.fill = blue_fill

    # 2. Print Remaining Standalone Results
    for category_key in ['bom', 'image', 'doc']:
        for risk_level, items in all_results.get(category_key, {}).items():
            for item in items:
                if category_key == 'bom':
                    row = ["BOM", item.get('Risk Level'), item.get('Category'), item.get('Product'), item.get('Company'),
                           item.get('Finding')]
                elif category_key == 'image':
                    finding = item.get('Finding', '')
                    if item.get('Risk Level', 'low').lower() != 'low':
                        finding += f" (Match: {item.get('Brand Image', 'N/A')})"
                    row = ["Image", item.get('Risk Level'), item.get('Category'), item.get('Uploaded Image'),
                           item.get('Company'), finding]
                else:
                    continue
                ws.append(row)
                for cell in ws[ws.max_row]: cell.fill = get_risk_fill(item.get('Risk Level'))

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        for cell in col:
            try:
                if len(str(cell.value)) > max_length: max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[col[0].column_letter].width = (max_length + 2)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    current_date = datetime.datetime.now().strftime("%Y-%m-%d")
    return Response(output.getvalue(), mimetype="application/vnd.openxmlformats-officedocument-spreadsheetml-sheet",
                    headers={"Content-Disposition": f"attachment;filename=ip_risk_report_{current_date}.xlsx"})


if __name__ == '__main__':
    app.run(debug=True, port=5001)
